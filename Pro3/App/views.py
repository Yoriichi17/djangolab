from django.shortcuts import render

from .models import Publisher

# Create your views here.

def fetch_publisher_data():
    
    publisher_data = Publisher.objects.all().values_list('name','address','city','state_province','country','website')
    
    return ['Publisher Name','Publisher Address','Publisher City','Publisher State province','Publisher Country','Publisher Website'],list(publisher_data)



from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table
from io import BytesIO

def create_pdf(column_names, students_data):
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    
    table_data = [column_names] + students_data
    table = Table(table_data)
    pdf.build([table])
    
    buffer.seek(0)
    return buffer

from django.http import FileResponse

def download_students_pdf(request):
    column_names, students_data = fetch_publisher_data()
    pdf_buffer = create_pdf(column_names, students_data)
    
    return FileResponse(pdf_buffer, as_attachment=True, filename='students_data.pdf')


import openpyxl
from io import BytesIO

def create_excel(column_names, students_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

  
    for col_num, column_title in enumerate(column_names, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title


    for row_num, row_data in enumerate(students_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value


    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def download_students_excel(request):
    column_names, students_data = fetch_publisher_data()
    excel_buffer = create_excel(column_names, students_data)
    
    return FileResponse(excel_buffer, as_attachment=True, filename='students_data.xlsx')

